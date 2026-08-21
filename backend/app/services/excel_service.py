"""档案 Excel 导入导出工具。"""
from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook, load_workbook

HEADERS = [
    "档案名称",
    "分类",
    "颁发机构",
    "颁发日期",
    "有效期至",
    "证书编号",
    "等级/成绩",
    "持有人",
    "标签",
    "关联经历",
    "备注",
]


def _fmt_date(value: date | datetime | None) -> str:
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return value.strftime("%Y-%m-%d")


def parse_date_cell(value) -> date | None:
    """解析 Excel 日期单元格，失败返回 None。"""
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _row_values(archive) -> list:
    tags = "、".join(t.name for t in archive.tags) if archive.tags else ""
    return [
        archive.name,
        archive.category.name if archive.category else "",
        archive.issuer or "",
        _fmt_date(archive.issue_date),
        _fmt_date(archive.expire_date),
        archive.cert_no or "",
        archive.grade or "",
        archive.holder or "",
        tags,
        archive.related_experience or "",
        archive.remark or "",
    ]


def build_export_workbook(archives) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "档案清单"
    ws.append(HEADERS)
    for archive in archives:
        ws.append(_row_values(archive))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_template_workbook() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "档案导入模板"
    ws.append(HEADERS)
    ws.append(
        [
            "示例：大学英语六级证书",
            "获奖证书类",
            "教育部考试中心",
            "2024-06-01",
            "2028-06-01",
            "CET6-123456",
            "520分",
            "张三",
            "英语,证书",
            "大学期间取得",
            "示例备注",
        ]
    )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def parse_import_rows(content: bytes) -> list[dict]:
    wb = load_workbook(BytesIO(content))
    ws = wb.active
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        rows.append(dict(zip(HEADERS, row)))
    return rows
